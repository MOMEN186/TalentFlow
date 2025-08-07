from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Employee, LeaveNote,PayRoll
from .serializers import EmployeeSerializer, LeaveNoteSerializer,PayRollSerializer
from django.utils.timezone import now
import openpyxl


class EmployeeViewSet(viewsets.ModelViewSet):

    queryset = Employee.objects.all().select_related(
        "department",
        "job_title",
    )
    serializer_class = EmployeeSerializer

    # Override retrieve to return custom key
    def retrieve(self, request, pk=None):
        """
        GET /employees/{pk}/
        Returns single employee data under 'employee' key.
        """
        employee = self.get_object()
        serializer = self.get_serializer(employee)
        return Response({"employee": serializer.data}, status=status.HTTP_200_OK)

    

class LeaveNoteViewSet(viewsets.ModelViewSet):
    queryset = LeaveNote.objects.all().select_related("employee")
    serializer_class = LeaveNoteSerializer
    
    @action(detail=False, methods=["get"], url_path="leave_notes")
    def leave_notes(self, request):
        query = LeaveNote.objects.select_related("employee").filter(
            date__gt=now().date()
        )
        serializer = LeaveNoteSerializer(query, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="export")
    def leave_notes_export(self, request):
        """
        GET /api/employees/leave_notes/export/
        • ?excel=true → streams an .xlsx file of future leave notes
        • otherwise   → returns JSON of future leave notes
        """
        # 1) only future notes
        future_notes = LeaveNote.objects.select_related("employee").filter(
            date__gt=now().date()
        )

        # 2) Excel branch?
        if request.query_params.get("excel") in ("1", "true", "yes"):
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="LeaveNotes")
            ws.append(["ID", "Employee", "From", "To", "Reason"])
            for note in future_notes:
                full_name = " ".join(
                    filter(
                        None,
                        [
                            note.employee.first_name,
                            note.employee.middle_name,
                            note.employee.last_name,
                        ],
                    )
                )
                ws.append(
                    [
                        note.id,
                        full_name,
                        note.date.strftime("%Y-%m-%d"),
                        note.return_date.strftime("%Y-%m-%d"),
                        note.description,
                    ]
                )
            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="leave_notes.xlsx"'
            wb.save(resp)
            return resp

        # 3) JSON branch
        serializer = LeaveNoteSerializer(future_notes, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





class PayRollViewSet(viewsets.ModelViewSet):
    queryset = PayRoll.objects.select_related("employee").all()
    serializer_class = PayRollSerializer

    
    @action(detail=False, methods=["get"], url_path="")
    def payroll(self, request):
        """
        GET /employees/payroll/
        Returns all employees with related salary, department, and job_title.
        """
      
        qs = self.get_queryset()
        serializer = self.get_serializer(qs, many=True)
        return Response({"payroll": serializer.data}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="export")
    def payroll_export(self, request):
        # Fetch all employees with related data
        employees = self.get_queryset()

        # If ?excel=true, generate and return .xlsx
        if request.query_params.get("excel") in ("1", "true", "yes"):
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="Payroll")

            # Header row
            ws.append([
                "ID",
                "First Name",
                "Middle Name",
                "Last Name",
                "Email",
                "Phone",
                "Department",
                "Job Title",
                "Gross Pay",
                "Tax",
                "Bonus",
                "Net Pay",
            ])

            # Data rows
            for emp in employees:
                ws.append([
                    emp.id,
                    emp.first_name,
                    emp.middle_name or "",
                    emp.last_name,
                    emp.email,
                    emp.phone,
                    emp.department.name if emp.department else "",
                    emp.job_title.name if emp.job_title else "",
                    emp.salary.gross_pay,
                    emp.salary.tax,
                    emp.salary.bonus,
                    emp.salary.net_pay,
                ])

            # Stream workbook to response
            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="payroll.xlsx"'
            wb.save(resp)
            return resp

        # Otherwise fall back to JSON
        serializer = self.get_serializer(employees, many=True)
        return Response({"payroll": serializer.data}, status=status.HTTP_200_OK)